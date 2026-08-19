from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .food_library import IMPORT_FIELDS, detect_mappings, parse_rows

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 10000


@dataclass(frozen=True)
class ParsedImport:
    source_type: str
    source_name: str
    sheet_names: list[str]
    selected_sheet: str | None
    headers: list[str]
    mappings: dict[str, str | None]
    rows: list[dict[str, Any]]


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV must be UTF-8 encoded")


def parse_csv_bytes(data: bytes, filename: str = "upload.csv", mappings: dict[str, str] | None = None) -> ParsedImport:
    if not data or len(data) > MAX_IMPORT_BYTES:
        raise ValueError("CSV is empty or exceeds the 5 MB upload limit")
    text = _decode_csv(data)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = list(reader.fieldnames or [])
    if not headers:
        raise ValueError("CSV header row could not be detected")
    raw_rows = list(reader)
    if len(raw_rows) > MAX_IMPORT_ROWS:
        raise ValueError(f"CSV exceeds the {MAX_IMPORT_ROWS} row import limit")
    detected = detect_mappings(headers)
    if mappings:
        detected.update({field: header for field, header in mappings.items() if field in IMPORT_FIELDS})
    return ParsedImport("csv", Path(filename).name, [], None, headers, detected, parse_rows(raw_rows, detected))


def _worksheet_rows(sheet) -> tuple[list[str], list[dict[str, Any]]]:
    values = list(sheet.iter_rows(values_only=True))
    values = [row for row in values if any(value is not None and str(value).strip() for value in row)]
    if not values:
        return [], []
    headers = [str(value or "").strip() for value in values[0]]
    rows: list[dict[str, Any]] = []
    for values_row in values[1:MAX_IMPORT_ROWS + 1]:
        rows.append({headers[index]: values_row[index] if index < len(values_row) else None for index in range(len(headers))})
    return headers, rows


def list_xlsx_sheets(data: bytes) -> list[str]:
    if not data or len(data) > MAX_IMPORT_BYTES:
        raise ValueError("XLSX is empty or exceeds the 5 MB upload limit")
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    names: list[str] = []
    for name in workbook.sheetnames:
        headers, rows = _worksheet_rows(workbook[name])
        if headers or rows:
            names.append(name)
    workbook.close()
    return names


def parse_xlsx_bytes(data: bytes, filename: str = "upload.xlsx", sheet_name: str | None = None, mappings: dict[str, str] | None = None) -> ParsedImport:
    sheets = list_xlsx_sheets(data)
    if not sheets:
        raise ValueError("Workbook does not contain a non-empty worksheet")
    selected = sheet_name or sheets[0]
    if selected not in sheets:
        raise ValueError("Selected worksheet is empty or does not exist")
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    headers, raw_rows = _worksheet_rows(workbook[selected])
    workbook.close()
    if len(raw_rows) > MAX_IMPORT_ROWS:
        raise ValueError(f"Worksheet exceeds the {MAX_IMPORT_ROWS} row import limit")
    detected = detect_mappings(headers)
    if mappings:
        detected.update({field: header for field, header in mappings.items() if field in IMPORT_FIELDS})
    return ParsedImport("xlsx", Path(filename).name, sheets, selected, headers, detected, parse_rows(raw_rows, detected))


def template_csv() -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(IMPORT_FIELDS)
    writer.writerow(["Example cereal", "Example brand", "Breakfast cereal", "40", "g", "per_100g", "100", "g", "370", "9", "70", "4", "1", "18", "8", "450", "120", "8", "300", "0", "0", "Example only"])
    return output.getvalue()


def template_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Foods"
    sheet.append(IMPORT_FIELDS)
    sheet.append(["Example cereal", "Example brand", "Breakfast cereal", 40, "g", "per_100g", 100, "g", 370, 9, 70, 4, 1, 18, 8, 450, 120, 8, 300, 0, 0, "Example only"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
